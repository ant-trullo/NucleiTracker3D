"""This function save information about nuclei intensity.

Input are the matrix of segmented and tracked nuclei and raw data in 2 channels.
For each of the tracked nucleus, the function writes an Excel file with the intensity in both channels.
"""

import os
import datetime
import time
import numpy as np
import xlsxwriter
from openpyxl import load_workbook
from skimage.measure import regionprops_table
from skimage.segmentation import expand_labels

import ServiceWidgets


class AnalysisSaver:
    """Only class, does all the job."""
    def __init__(self, nucs_segm, nucs_trck, raw_data, folder2write, chs_red_green, soft_version, fname4journal):

        nucs_tags  =  np.unique(nucs_trck[nucs_trck != 0])
        tlen       =  nucs_trck.shape[0]

        green_ints_vol  =  np.zeros((nucs_tags.size, nucs_trck.shape[0], 2))
        red_ints_vol    =  np.zeros((nucs_tags.size, nucs_trck.shape[0], 2))

        for tt in range(tlen):
            rgp_red_bff  =  regionprops_table(nucs_trck[tt], raw_data.red[tt], properties=["label", "area", "image_intensity"])
            for cntr, rr in enumerate(rgp_red_bff["label"]):
                ll                    =  np.where(nucs_tags == rr)[0][0]
                red_ints_vol[ll, tt]  =  np.sum(rgp_red_bff["image_intensity"][cntr]), rgp_red_bff["area"][cntr]

            rgp_green_bff  =  regionprops_table(nucs_trck[tt], raw_data.green[tt], properties=["label", "area", "image_intensity"])
            for cntg, gg in enumerate(rgp_green_bff["label"]):
                oo                      =  np.where(nucs_tags == gg)[0][0]
                green_ints_vol[oo, tt]  =  np.sum(rgp_green_bff["image_intensity"][cntg]), rgp_green_bff["area"][cntg]

        workbook          =  xlsxwriter.Workbook(folder2write + "/NucleiIntensity.xlsx")
        sheet_info        =  workbook.add_worksheet("Info")
        sheet_red_ints    =  workbook.add_worksheet("Red Ints")
        sheet_red_vols    =  workbook.add_worksheet("Red Vols")
        sheet_green_ints  =  workbook.add_worksheet("Green Ints")
        sheet_green_vols  =  workbook.add_worksheet("Green Vols")

        sheet_info.write(0, 0, "date")
        sheet_info.write(0, 1, datetime.date.today().strftime("%d%b%Y"))
        sheet_info.write(1, 0, "software version")
        sheet_info.write(1, 1, soft_version)
        sheet_info.write(2, 0, "x-y pixel size µm")
        sheet_info.write(2, 1, raw_data.pix_size_xy)
        sheet_info.write(3, 0, "z pixel size in µm")
        sheet_info.write(3, 1, raw_data.pix_size_z)
        sheet_info.write(4, 0, "time step in s")
        sheet_info.write(4, 1, raw_data.time_step_value)
        # sheet_info.write(5, 0, "distance threshold")
        # sheet_info.write(5, 1, dist_thr)
        sheet_info.write(6, 0, "Raw Filenames")
        for ccc, fname in enumerate(fname4journal):
            sheet_info.write(ccc + 6, 1, fname)

        sheet_red_ints.write(0, 0, "Time")
        sheet_red_ints.write(0, 1, "Frame")
        sheet_red_vols.write(0, 0, "Time")
        sheet_red_vols.write(0, 1, "Frame")
        sheet_green_ints.write(0, 0, "Time")
        sheet_green_ints.write(0, 1, "Frame")
        sheet_green_vols.write(0, 0, "Time")
        sheet_green_vols.write(0, 1, "Frame")

        for dd in range(tlen):
            sheet_red_ints.write(1 + dd, 1, dd)
            sheet_red_vols.write(1 + dd, 1, dd)
            sheet_green_ints.write(1 + dd, 1, dd)
            sheet_green_vols.write(1 + dd, 1, dd)

            sheet_red_ints.write(1 + dd, 0, time.strftime("%M:%S", time.gmtime(dd * raw_data.time_step_value)))
            sheet_red_vols.write(1 + dd, 0, time.strftime("%M:%S", time.gmtime(dd * raw_data.time_step_value)))
            sheet_green_ints.write(1 + dd, 0, time.strftime("%M:%S", time.gmtime(dd * raw_data.time_step_value)))
            sheet_green_vols.write(1 + dd, 0, time.strftime("%M:%S", time.gmtime(dd * raw_data.time_step_value)))

        for cnt_t, nuc_tag in enumerate(nucs_tags):
            sheet_red_ints.write(0, 2 + cnt_t, "Nuc_" + str(nuc_tag))
            sheet_red_vols.write(0, 2 + cnt_t, "Nuc_" + str(nuc_tag))
            sheet_green_ints.write(0, 2 + cnt_t, "Nuc_" + str(nuc_tag))
            sheet_green_vols.write(0, 2 + cnt_t, "Nuc_" + str(nuc_tag))

        for ll_r in range(red_ints_vol.shape[0]):
            for tt_r in range(tlen):
                sheet_red_ints.write(1 + tt_r, 2 + ll_r, red_ints_vol[ll_r, tt_r, 0])
                sheet_red_vols.write(1 + tt_r, 2 + ll_r, red_ints_vol[ll_r, tt_r, 1])

        for ll_g in range(green_ints_vol.shape[0]):
            for tt_g in range(tlen):
                sheet_green_ints.write(1 + tt_g, 2 + ll_g, green_ints_vol[ll_g, tt_g, 0])
                sheet_green_vols.write(1 + tt_g, 2 + ll_g, green_ints_vol[ll_g, tt_g, 1])

        workbook.close()

        np.save(folder2write + '/first_slctd_frame.npy', raw_data.red_mip[0])
        np.save(folder2write + '/last_slctd_frame.npy', raw_data.red_mip[-1])
        np.save(folder2write + '/nucs_segm.npy', nucs_segm)
        np.save(folder2write + '/nucs_trck.npy', nucs_trck)
        np.save(folder2write + '/chs_red_green.npy', chs_red_green)
        np.save(folder2write +  '/red_ints_vol.npy', red_ints_vol)
        np.save(folder2write +  '/green_ints_vol.npy', green_ints_vol)


class ShowTraces:
    """Collect intensity and volume of a tracked 3D nucleus."""
    def __init__(self, nucs_trck, nucs_tag, raw_data):

        tlen            =  nucs_trck.shape[0]                                                                           # number of time steps
        red_ints_vol    =  np.zeros((tlen, 2))                                                                          # initialize output matrix with intensity and volume in the red channel
        green_ints_vol  =  np.zeros((tlen, 2))                                                                          # initialize output matrix with intensity and volume in the green channel

        for tt in range(tlen):                                                                                          # for each time frame
            rgp_red_bff            =  regionprops_table(nucs_trck[tt], raw_data.red[tt], properties=["label", "area", "image_intensity"])       # regionprops on the 3D nuclei single time frame
            ll                     =  np.where(rgp_red_bff["label"] == nucs_tag)[0][0]                                                          # coordinate of the selected nucleus
            red_ints_vol[tt, :]    =  np.sum(rgp_red_bff["image_intensity"][ll]), rgp_red_bff["area"][ll]                                       # insert the total intensity of the nucleus and its volume
            rgp_green_bff          =  regionprops_table(nucs_trck[tt], raw_data.green[tt], properties=["label", "area", "image_intensity"])     # same for the green
            oo                     =  np.where(rgp_green_bff["label"] == nucs_tag)[0][0]
            green_ints_vol[tt, :]  =  np.sum(rgp_green_bff["image_intensity"][oo]), rgp_green_bff["area"][oo]

        self.red_avints    =  red_ints_vol[:, 0] / red_ints_vol[:, 1]                                                   # average nucleus intensity as output (for both channels)
        self.green_avints  =  green_ints_vol[:, 0] / green_ints_vol[:, 1]


class SavePatchPostProcessing:
    """This function is used to correct a problem with our first analysis."""
    def __init__(self, raw_data, analysis_folder, y_coord, xlsx2correct, soft_version, fname4journal):

        book2correct   =  load_workbook(xlsx2correct)
        sheet_inpatt   =  book2correct["Inside Pattern"]
        sheet_outpatt  =  book2correct["Outside Pattern"]

        av_ints_inpatt   =  []
        av_ints_outpatt  =  []
        for mm in range(1, sheet_inpatt.max_row):
            av_ints_inpatt.append(sheet_inpatt.cell(mm + 1, 4).value)
            av_ints_outpatt.append(sheet_outpatt.cell(mm + 1, 4).value)

        nucs_segm  =  np.load(analysis_folder + '/nucs_segm.npy')
        tlen       =  nucs_segm.shape[0]

        bkg_in_nucs   =  []
        bkg_out_nucs  =  []
        bkg_evwhere   =  []
        for tt in range(tlen):
            nucs_segm_bff                  =  np.sign(nucs_segm[tt])
            nucs_segm_bff[:, :, y_coord:]  =  0
            in_nucs                        =  raw_data.green[tt] * nucs_segm_bff
            if in_nucs.sum() == 0 or nucs_segm_bff.sum() == 0:
                bkg_in_nucs.append(0)
            else:
                bkg_in_nucs.append(in_nucs.sum() / nucs_segm_bff.sum())
            # bkg_in_nucs.append(in_nucs.sum() / nucs_segm_bff.sum())

            nucs_segm_bff                  =  1 - np.sign(nucs_segm[tt])
            nucs_segm_bff[:, :, y_coord:]  =  0
            out_nucs                       =  raw_data.green[tt] * nucs_segm_bff
            bkg_out_nucs.append(out_nucs.sum() / nucs_segm_bff.sum())

            nucs_segm_bff                  =  np.ones_like(nucs_segm[tt])
            nucs_segm_bff[:, :, y_coord:]  =  0
            evwhere_nucs                   =  raw_data.green[tt] * nucs_segm_bff
            bkg_evwhere.append(evwhere_nucs.sum() / nucs_segm_bff.sum())

        tot_bkg_av  =  []
        for ll in range(tlen):
            msk_bff  =  expand_labels((1 - np.sign(nucs_segm[ll])))
            tot_bkg_av.append((msk_bff * raw_data.green[ll]).sum() / msk_bff.sum())
        tot_bkg_av  =  np.asarray(tot_bkg_av).mean()

        workbook    =  xlsxwriter.Workbook(analysis_folder + "/NucleiIntensityBackground.xlsx")
        sheet_info  =  workbook.add_worksheet("Info")
        sheet1      =  workbook.add_worksheet("Background")
        sheet2      =  workbook.add_worksheet("In pattern")
        sheet3      =  workbook.add_worksheet("Out pattern")

        sheet_info.write(0, 0, "date")
        sheet_info.write(0, 1, datetime.date.today().strftime("%d%b%Y"))
        sheet_info.write(1, 0, "software version")
        sheet_info.write(1, 1, soft_version)
        sheet_info.write(2, 0, "x-y pixel size µm")
        sheet_info.write(2, 1, raw_data.pix_size_xy)
        sheet_info.write(3, 0, "z pixel size in µm")
        sheet_info.write(3, 1, raw_data.pix_size_z)
        sheet_info.write(4, 0, "time step in s")
        sheet_info.write(4, 1, raw_data.time_step_value)
        sheet_info.write(6, 0, "Raw Filenames")
        for ccc, fname in enumerate(fname4journal):
            sheet_info.write(ccc + 6, 1, fname)

        sheet1.write(0, 0, "Time Frame")
        sheet2.write(0, 0, "Time Frame")
        sheet3.write(0, 0, "Time Frame")
        sheet1.write(0, 1, "Time")
        sheet2.write(0, 1, "Time")
        sheet3.write(0, 1, "Time")
        for uu in range(tlen):
            sheet1.write(uu + 1, 0, uu)
            sheet2.write(uu + 1, 0, uu)
            sheet3.write(uu + 1, 0, uu)
            sheet1.write(uu + 1, 1, uu * raw_data.time_step_value)
            sheet2.write(uu + 1, 1, uu * raw_data.time_step_value)
            sheet3.write(uu + 1, 1, uu * raw_data.time_step_value)

        sheet1.write(0, 2, "Bckg_InNucs")
        sheet1.write(0, 3, "Bckg_OutNucs")
        sheet1.write(0, 4, "Bckg_EvWhere")

        sheet2.write(0, 2, "Av Ints")
        sheet2.write(0, 3, "AvInts - Bckg_InNucs")
        sheet2.write(0, 4, "AvInts - Bckg_OutNucs")
        sheet2.write(0, 5, "AvInts - Bckg_EvWhere")
        sheet2.write(0, 6, "(AvInts - Bckg_InNucs)/ tot_bkg")
        sheet2.write(0, 7, "(AvInts - Bckg_OutNucs)/ tot_bkg")
        sheet2.write(0, 8, "(AvInts - Bckg_EvWhere)/ tot_bkg")

        sheet3.write(0, 2, "Av Ints")
        sheet3.write(0, 3, "AvInts - Bckg_InNucs")
        sheet3.write(0, 4, "AvInts - Bckg_OutNucs")
        sheet3.write(0, 5, "AvInts - Bckg_EvWhere")
        sheet3.write(0, 6, "(AvInts - Bckg_InNucs)/ tot_bkg")
        sheet3.write(0, 7, "(AvInts - Bckg_OutNucs)/ tot_bkg")
        sheet3.write(0, 8, "(AvInts - Bckg_EvWhere)/ tot_bkg")

        for oo in range(tlen):
            sheet1.write(oo + 1, 2, bkg_in_nucs[oo])
            sheet1.write(oo + 1, 3, bkg_out_nucs[oo])
            sheet1.write(oo + 1, 4, bkg_evwhere[oo])

            sheet2.write(oo + 1, 2, av_ints_inpatt[oo])
            sheet2.write(oo + 1, 3, av_ints_inpatt[oo] - bkg_in_nucs[oo])
            sheet2.write(oo + 1, 4, av_ints_inpatt[oo] - bkg_out_nucs[oo])
            sheet2.write(oo + 1, 5, av_ints_inpatt[oo] - bkg_evwhere[oo])
            sheet2.write(oo + 1, 6, (av_ints_inpatt[oo] - bkg_in_nucs[oo]) / tot_bkg_av)
            sheet2.write(oo + 1, 7, (av_ints_inpatt[oo] - bkg_out_nucs[oo]) / tot_bkg_av)
            sheet2.write(oo + 1, 8, (av_ints_inpatt[oo] - bkg_evwhere[oo]) / tot_bkg_av)

            sheet3.write(oo + 1, 2, av_ints_outpatt[oo])
            sheet3.write(oo + 1, 3, av_ints_outpatt[oo] - bkg_in_nucs[oo])
            sheet3.write(oo + 1, 4, av_ints_outpatt[oo] - bkg_out_nucs[oo])
            sheet3.write(oo + 1, 5, av_ints_outpatt[oo] - bkg_evwhere[oo])
            sheet3.write(oo + 1, 6, (av_ints_outpatt[oo] - bkg_in_nucs[oo]) / tot_bkg_av)
            sheet3.write(oo + 1, 7, (av_ints_outpatt[oo] - bkg_out_nucs[oo]) / tot_bkg_av)
            sheet3.write(oo + 1, 8, (av_ints_outpatt[oo] - bkg_evwhere[oo]) / tot_bkg_av)

        sheet1.write(0, 7, "Total Background")
        sheet1.write(0, 8, tot_bkg_av)

        workbook.close()


class TracesSpatialWithBackground:
    """Only class, does all the job"""
    def __init__(self, analysis_folder, raw_data, nucs3d_fin, roi_line, soft_version, fname4journal):

        y_coord   =  int(roi_line.getRegion()[1])
        tlen      =  nucs3d_fin.shape[0]
        prof_in   =  np.zeros((tlen, 2))                                                                                # profile intensity and volume inside the pattern
        prof_out  =  np.zeros((tlen, 2))                                                                                # profile intensity and volume outside the pattern

        pbar  =  ServiceWidgets.ProgressBar(total1=tlen)
        pbar.update_progressbar1(0)
        pbar.show()

        for tt in range(tlen):
            # print(tt)
            pbar.update_progressbar1(tt)
            rgp_bff  =  regionprops_table(nucs3d_fin[tt], raw_data.green[tt], properties=["label", "intensity_image", "centroid", "area"])
            for uu in range(len(rgp_bff["area"])):
                if rgp_bff["centroid-2"][uu] > y_coord:
                    prof_in[tt, :]  +=  np.sum(rgp_bff["intensity_image"][uu]), rgp_bff["area"][uu]
                    # aze_in[tt]  +=  (nucs3d_fin[tt] == rgp_bff["label"][uu]).sum(0).astype(np.uint16)
                elif rgp_bff["centroid-2"][uu] <= y_coord:
                    prof_out[tt, :]  +=  np.sum(rgp_bff["intensity_image"][uu]), rgp_bff["area"][uu]
                    # aze_out[tt]  +=  (nucs3d_fin[tt] == rgp_bff["label"][uu]).sum(0).astype(np.uint16)

        self.prof_in   =  prof_in
        self.prof_out  =  prof_out

        np.save(analysis_folder + '/prof_in.npy', prof_in)
        np.save(analysis_folder + '/prof_out.npy', prof_out)

        book2wrt    =  xlsxwriter.Workbook(analysis_folder + '/SpatialTracesWithBackground.xlsx')                                                                  # write results
        sheet_info  =  book2wrt.add_worksheet("Info")
        sheet1      =  book2wrt.add_worksheet("Inside Pattern")
        sheet2      =  book2wrt.add_worksheet("Outside Pattern")
        sheet3      =  book2wrt.add_worksheet("Background")

        sheet_info.write(0, 0, "date")
        sheet_info.write(0, 1, datetime.date.today().strftime("%d%b%Y"))
        sheet_info.write(1, 0, "software version")
        sheet_info.write(1, 1, soft_version)
        sheet_info.write(2, 0, "x-y pixel size µm")
        sheet_info.write(2, 1, raw_data.pix_size_xy)
        sheet_info.write(3, 0, "z pixel size in µm")
        sheet_info.write(3, 1, raw_data.pix_size_z)
        sheet_info.write(4, 0, "time step in s")
        sheet_info.write(4, 1, raw_data.time_step_value)
        sheet_info.write(5, 0, "Line position")
        sheet_info.write(5, 1, y_coord * raw_data.pix_size_xy)
        sheet_info.write(7, 0, "Raw Filenames")
        for ccc, fname in enumerate(fname4journal):
            sheet_info.write(ccc + 7, 1, fname)

        sheet1.write(0, 0, "Time Step")
        sheet1.write(0, 1, "Tot Intensity")
        sheet1.write(0, 2, "Volume")
        sheet1.write(0, 3, "Av Intensity")
        sheet1.write(0, 4, "AvInts - Bckg_InNucs")
        sheet1.write(0, 5, "AvInts - Bckg_OutNucs")
        sheet1.write(0, 6, "AvInts - Bckg_EvWhere")
        sheet1.write(0, 7, "(AvInts - Bckg_InNucs) / tot_bkg")
        sheet1.write(0, 8, "(AvInts - Bckg_OutNucs) / tot_bkg")
        sheet1.write(0, 9, "(AvInts - Bckg_EvWhere) / tot_bkg")

        sheet2.write(0, 0, "Time Step")
        sheet2.write(0, 1, "Tot Intensity")
        sheet2.write(0, 2, "Volume")
        sheet2.write(0, 3, "Av Intensity")
        sheet2.write(0, 4, "AvInts - Bckg_InNucs")
        sheet2.write(0, 5, "AvInts - Bckg_OutNucs")
        sheet2.write(0, 6, "AvInts - Bckg_EvWhere")
        sheet2.write(0, 7, "(AvInts - Bckg_InNucs) / tot_bkg")
        sheet2.write(0, 8, "(AvInts - Bckg_OutNucs) / tot_bkg")
        sheet2.write(0, 9, "(AvInts - Bckg_EvWhere) / tot_bkg")

        for cnt, ll in enumerate(prof_in):
            sheet1.write(cnt + 1, 0, cnt)
            sheet3.write(cnt + 1, 0, cnt)
            sheet1.write(cnt + 1, 1, ll[0])
            sheet1.write(cnt + 1, 2, ll[1])
            sheet1.write(cnt + 1, 3, ll[0] / ll[1])

        for tnc, jj in enumerate(prof_out):
            sheet2.write(tnc + 1, 0, tnc)
            sheet2.write(tnc + 1, 1, jj[0])
            sheet2.write(tnc + 1, 2, jj[1])
            sheet2.write(tnc + 1, 3, jj[0] / jj[1])

        tlen          =  nucs3d_fin.shape[0]
        bkg_in_nucs   =  []
        bkg_out_nucs  =  []
        bkg_evwhere   =  []

        for tt in range(tlen):
            nucs_segm_bff                  =  np.sign(nucs3d_fin[tt])
            nucs_segm_bff[:, :, y_coord:]  =  0
            in_nucs                        =  raw_data.green[tt] * nucs_segm_bff
            bkg_in_nucs.append(in_nucs.sum() / nucs_segm_bff.sum())

            nucs_segm_bff                  =  1 - np.sign(nucs3d_fin[tt])
            nucs_segm_bff[:, :, y_coord:]  =  0
            out_nucs                       =  raw_data.green[tt] * nucs_segm_bff
            bkg_out_nucs.append(out_nucs.sum() / nucs_segm_bff.sum())

            nucs_segm_bff                  =  np.ones_like(nucs3d_fin[tt])
            nucs_segm_bff[:, :, y_coord:]  =  0
            evwhere_nucs                   =  raw_data.green[tt] * nucs_segm_bff
            bkg_evwhere.append(evwhere_nucs.sum() / nucs_segm_bff.sum())

        tot_bkg_av  =  []
        for ll in range(tlen):
            msk_bff  =  expand_labels((1 - np.sign(nucs3d_fin[ll])))
            tot_bkg_av.append((msk_bff * raw_data.green[ll]).sum() / msk_bff.sum())
        tot_bkg_av  =  np.asarray(tot_bkg_av).mean()

        sheet3.write(0, 0, "Time Frame")
        sheet3.write(0, 1, "Bckg_InNucs")
        sheet3.write(0, 2, "Bckg_OutNucs")
        sheet3.write(0, 3, "Bckg_EvWhere")

        av_ints_inpatt   =  prof_in[:, 0] / prof_in[:, 1]
        av_ints_outpatt  =  prof_out[:, 0] / prof_out[:, 1]
        for oo in range(tlen):
            sheet3.write(oo + 1, 1, bkg_in_nucs[oo])
            sheet3.write(oo + 1, 2, bkg_out_nucs[oo])
            sheet3.write(oo + 1, 3, bkg_evwhere[oo])

            sheet1.write(oo + 1, 4, av_ints_inpatt[oo] - bkg_in_nucs[oo])
            sheet1.write(oo + 1, 5, av_ints_inpatt[oo] - bkg_out_nucs[oo])
            sheet1.write(oo + 1, 6, av_ints_inpatt[oo] - bkg_evwhere[oo])
            sheet1.write(oo + 1, 7, (av_ints_inpatt[oo] - bkg_in_nucs[oo]) / tot_bkg_av)
            sheet1.write(oo + 1, 8, (av_ints_inpatt[oo] - bkg_out_nucs[oo]) / tot_bkg_av)
            sheet1.write(oo + 1, 9, (av_ints_inpatt[oo] - bkg_evwhere[oo]) / tot_bkg_av)

            sheet2.write(oo + 1, 4, av_ints_outpatt[oo] - bkg_in_nucs[oo])
            sheet2.write(oo + 1, 5, av_ints_outpatt[oo] - bkg_out_nucs[oo])
            sheet2.write(oo + 1, 6, av_ints_outpatt[oo] - bkg_evwhere[oo])
            sheet2.write(oo + 1, 7, (av_ints_outpatt[oo] - bkg_in_nucs[oo]) / tot_bkg_av)
            sheet2.write(oo + 1, 8, (av_ints_outpatt[oo] - bkg_out_nucs[oo]) / tot_bkg_av)
            sheet2.write(oo + 1, 9, (av_ints_outpatt[oo] - bkg_evwhere[oo]) / tot_bkg_av)

        sheet3.write(0, 7, "Total Background")
        sheet3.write(0, 8, tot_bkg_av)

        book2wrt.close()
